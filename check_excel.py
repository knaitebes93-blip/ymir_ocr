#!/usr/bin/env python3
import openpyxl

wb = openpyxl.load_workbook('precios_market.xlsx')
print(f'Hojas: {wb.sheetnames}\n')

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'=== {sheet_name.upper()} ===')
    for i, row in enumerate(ws.iter_rows(max_row=15), 1):
        print(f'Row {i}: {row[0].value} | {row[1].value} | {row[2].value} | {row[3].value}')
    print()

wb.close()
